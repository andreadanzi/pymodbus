# -*- coding: utf-8 -*-
#!/usr/bin/env python
import pymongo
from pymongo import MongoClient
from bson.objectid import ObjectId
import logging
import logging.handlers
import os
import sys
import getopt
import datetime
import collections
from scipy.stats import randint
import numpy as np
from collections import namedtuple
from openpyxl import *
StepRow = namedtuple('StepRow',['boreholeid','station','elevation',
                                'stage_sequence','method','top_depth',
                                'top_elevation','stage_length','mixtype',
                                'refusalP','minFlowRate','maxVolume','groutMix',
                                'startStepTime','stopStepTime','elapsedTime','productionTime',
                                'groutTake', 'flowRateAvg', 'flowRateMax', 'flowRateFinal',
                                 'pGaugeAvg', 'pGaugeMax','qAtpGaugeMax', 'pGaugeFinal','pEffAvg', 'pEffMax','qAtpEffMax', 'pEffFinal','appLugeonUnit','gin','pq'
                                ])
"""
	3	2	1	k
Q	0.0001447	-0.024499	0.86506	15.4763
P	0.000021167	-0.00403	0.30249	-0.8024
"""

bh_type_codes = {"E":"Exploratory","P":"Primary","S":"Secondary","T":"Ternary","Q":"Quanternary","L":"Quinary"}

def pOut(t):
    retV=0.000021167*t**3-0.00403*t**2+0.30249*t-0.8024
    if retV < 0.0:
        retV = 0.0
    return retV

def qOut(t):
    retV = 0.0001447*t**3-0.024499*t**2+0.86506*t+15.4763
    return retV

# client = MongoClient('localhost', 27017)
log = logging.getLogger()
log.setLevel(logging.DEBUG)
file_handler = logging.handlers.RotatingFileHandler("{0}.log".format(os.path.basename(__file__).split(".")[0]), maxBytes=5000000,backupCount=5)
file_handler.setLevel(logging.DEBUG)
formatter = logging.Formatter('%(asctime)s %(name)-12s %(levelname)-8s %(message)s')
file_handler.setFormatter(formatter)
log.addHandler(file_handler)

delta_rand = randint(-2, 2)
# dati di partenza
# da caricare con
# >mongoimport --db import --collection timeseries --host srv3.sws-digital.com --file import_series.json
# generati con
# >mongoexport --db tgrout-development --collection timeseries --host srv3.sws-digital.com --out import_series.json --query "{'stage':ObjectId('57c6cbb9ea71da9c2513d5b2'),'step':ObjectId('57c6cbb9ea71da9c2513d5b3')}"
# 57cd5367deeb0e074430c8f7
# 57cd5367deeb0e074430c8f8
import_refusalPressure = 10.
base_stageId = "57c6cbb9ea71da9c2513d5b2"
base_stepId = "57c6cbb9ea71da9c2513d5b3"

# python export_xlsx_steps.py -m localhost -p 27017 -d tgrout-dev -b 57ab48f1f194b0873319a12a
def main(argv):
    syntax = "python " + os.path.basename(__file__) + " -m <mongo host> -p <mongo port> -d <main database> -b <borehole_id>"
    mongo_host = "localhost"
    mongo_port = 27017
    mongo_database = "tgrout"
    borehole_id = None
    try:
        opts = getopt.getopt(argv, "hm:p:d:b:", ["mongohost=","port=","database=","borehole_id="])[0]
    except getopt.GetoptError:
        print syntax
        sys.exit(1)
    if len(opts) < 1:
        print syntax
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            print syntax
            sys.exit()
        elif opt in ("-p", "--port"):
            mongo_port = int(arg)
        elif opt in ("-m", "--mongohost"):
            mongo_host = arg
        elif opt in ("-d", "--database"):
            mongo_database = arg
        elif opt in ("-b", "--borehole_id"):
            borehole_id = arg
    all_boreholes = []
    mClient = MongoClient(mongo_host, mongo_port)
    db = mClient[mongo_database]
    if borehole_id:
        log.info("Trying connecting to {0}:{1}, database {2}".format(mongo_host,mongo_port,mongo_database))
        main_borehole = db.boreholes.find_one({"_id":ObjectId(borehole_id)})
        all_boreholes.append(main_borehole)
    else:
        all_boreholes = db.boreholes.find()
    for main_borehole in all_boreholes:
        log.info("Found borehole with Id {0}! topElevation_Build = {1}, topElevation_Build={2}, holeLength_Build={3} and holeLength_Build={4}".format(main_borehole["_id"],main_borehole["topElevation_Build"],main_borehole["topElevation_Build"],main_borehole["holeLength_Build"],main_borehole["holeLength_Build"]))
        stages = db.stages.find({"borehole":main_borehole["_id"]}).sort('startDateTime', pymongo.ASCENDING)
        bh_section = db.sections.find_one({"_id":main_borehole["section"]})
        bh_section_code = bh_section["code"]
        bh_line = db.lines.find_one({"_id":main_borehole["line"]})
        bh_line_code = bh_line["code"]
        print "line {1}, section {0}".format(bh_section_code,bh_line_code)
        # boreholeId boreholeId
        # section
        # line
        # type
        # position
        # stationId_Build
        # topElevation_Build
        # offset_Build
        # inclination_Build
        # azimuth_Build
        wb = load_workbook("xlsx/data_record_tmp.xlsx")        
        wsht = wb.get_sheet_by_name('Data record')
        export_stages = []
        stgs = list(stages)
        for stSeq, s in enumerate( stgs ):
            print "ID {0} stageStatus {1}".format(s["_id"],s["stageStatus"])
            log.info("ID {0} stageStatus {1}".format(s["_id"],s["stageStatus"]))
            cumV = 0.
            for ist, step in enumerate(s["steps"]):
                print "\tID {0} stageStatus {1}".format(step["_id"],step["stepStatus"])
                log.info("\tID {0} stageStatus {1}".format(step["_id"],step["stepStatus"]))
                timeseries = db.timeseries.find({"stage":ObjectId(s["_id"]),"step":ObjectId(step["_id"])}).sort('timestampMinute', pymongo.ASCENDING)
                lastPe = collections.deque(maxlen=2*60)
                lastPg = collections.deque(maxlen=2*60)
                lastQ = collections.deque(maxlen=2*60)
                avgPe = 0.
                avgPg = 0.
                avgQ = 0.
                avg3Pe = [0.,0.,0.]
                avg3Pg = [0.,0.,0.]
                avg3Q = [0.,0.,0.]
                maxPe = 0.
                maxPg = 0.
                maxQ = 0.
                qAtMaxPe = 0.
                qAtMaxPg = 0.
                iMaxAvg = 3
                
                iCount = 0
                if timeseries.count():
                    print "timeseries.count() = {}".format(timeseries.count())
                    startDateTime = timeseries[0]['timestampMinute']
                    for tv in list(timeseries):
                        print "timestampMinute {}".format(tv['timestampMinute'])
                        stopDateTime = tv['timestampMinute']
                        lastSeconds = 0
                        tvo = collections.OrderedDict(sorted(tv["values"].items(), key=lambda t: int(t[0])))
                        for vk in tvo:
                            lastSeconds += 1
                            iCount = iCount + 1
                            print iCount
                            """
                            "flowRateGaugeManifold" : 17.9745810337963,
                            "pressureGaugeManifold" : 0.625024111325,
                            "pressureEffective" : 0.520853426104166,
                            "elapsedTime" : 57000,
                            "_id" : ObjectId("57d2c23177c87c26d6a06971"),
                            "injectedGroutTakeVolume" : 15.9477824093287
                            """
                            lastQ.append(tvo[vk]["flowRateGaugeManifold"])
                            lastPe.append(tvo[vk]["pressureEffective"])
                            lastPg.append(tvo[vk]["pressureGaugeManifold"])
                            avgPe = avgPe + (tvo[vk]["pressureEffective"] - avgPe)/(iCount+1)
                            avgPg = avgPg + (tvo[vk]["pressureGaugeManifold"] - avgPg)/(iCount+1)
                            avgQ = avgQ + (tvo[vk]["flowRateGaugeManifold"] - avgQ)/(iCount+1)
                            avg3Pe[(iCount-1)%iMaxAvg] = tvo[vk]["pressureEffective"] 
                            avg3Pg[(iCount-1)%iMaxAvg] = tvo[vk]["pressureGaugeManifold"] 
                            avg3Q[(iCount-1)%iMaxAvg] = tvo[vk]["flowRateGaugeManifold"] 
                            if (iCount-1)%iMaxAvg == iMaxAvg-1:
                                if np.mean(avg3Pe) > maxPe:
                                    maxPe = np.mean(avg3Pe)
                                    qAtMaxPe = np.mean(avg3Q)
                                if np.mean(avg3Pg) > maxPg:
                                    maxPg = np.mean(avg3Pg)
                                    qAtMaxPg = np.mean(avg3Q)
                                if np.mean(avg3Q) > maxQ:
                                    maxQ = np.mean(avg3Q)
                    
                    stopDateTime = stopDateTime + datetime.timedelta(seconds=lastSeconds)                    
                    elapsedTime = stopDateTime - startDateTime
                    stageLength = s['bottomLength'] - s['topLength']
                    flowRateAvg = avgQ/stageLength
                    flowRateMax = maxQ/stageLength
                    flowRateFinal = np.mean(lastQ)/stageLength
                    pGaugeAvg = avgPg
                    pGaugeMax = maxPg
                    qAtpGaugeMax = qAtMaxPg/stageLength
                    pGaugeFinal = np.mean(lastPg)
                    pEffAvg = avgPe
                    pEffMax = maxPe
                    qAtpEffMax = qAtMaxPe/stageLength
                    pEffFinal = np.mean(lastPe)
                    appLugeonUnit = 10.*flowRateFinal/pEffFinal
                    gin = step['groutTake']*pEffFinal/stageLength
                    pq = pEffFinal/flowRateFinal
                    print "iCount = {}".format(iCount)
                    effTime = datetime.timedelta(seconds=iCount)
                    maxVolume = 99
                    headLossFactor = db.headlossfactors.find_one({'_id':ObjectId(step['headLossFactor'])})
                    groutMix = db.mixtypes.find_one({'_id':headLossFactor['mixType']})
                    cumV += step['groutTake']
                    #A21  AJ21                  
                    step_row = StepRow(main_borehole['boreholeId'], #A21+i
                                        main_borehole['stationId_Build'],
                                        main_borehole['topElevation_Build'],
                                        stSeq + 1,
                                        s['procedureType'],
                                        s['topLength'],
                                        s['topLength'],
                                        s['topLength'],
                                        step['mixTypeType'],
                                        s['refusalPressure'],
                                        s['minFlowRate'],
                                        maxVolume,
                                        groutMix['code'],
                                        str(startDateTime),
                                        str(stopDateTime),
                                        str(elapsedTime),
                                        str(effTime),
                                        step['groutTake'],
                                        flowRateAvg,
                                        flowRateMax,
                                        flowRateFinal,
                                        pGaugeAvg,
                                        pGaugeMax,
                                        qAtpGaugeMax,
                                        pGaugeFinal,
                                        pEffAvg,
                                        pEffMax,
                                        qAtpEffMax,
                                        pEffFinal,
                                        appLugeonUnit,
                                        gin,
                                        pq)
                    iStepRow = (stSeq+1)*(ist+1)
                    # Borehole data
                    _ = wsht.cell(column=1, row=iStepRow+20, value="%s" % main_borehole['boreholeId'])
                    _ = wsht.cell(column=2, row=iStepRow+20, value="%s" % main_borehole['stationId_Build'])
                    _ = wsht.cell(column=3, row=iStepRow+20, value= main_borehole['inclination_Build'])
                    _ = wsht.cell(column=4, row=iStepRow+20, value= main_borehole['topElevation_Build'])
                    # Stage data
                    _ = wsht.cell(column=5, row=iStepRow+20, value=stSeq + 1)
                    _ = wsht.cell(column=6, row=iStepRow+20, value="%s" % s['procedureType'])
                    _ = wsht.cell(column=7, row=iStepRow+20, value="%s" % "ND") #"Stage (top-down)"
                    topDepth = s['topLength']*np.cos(np.radians(main_borehole['inclination_Build']))
                    _ = wsht.cell(column=8, row=iStepRow+20, value=topDepth) # Top Depth
                    _ = wsht.cell(column=9, row=iStepRow+20, value=main_borehole['topElevation_Build']-topDepth)
                    _ = wsht.cell(column=10, row=iStepRow+20, value=stageLength)
                    _ = wsht.cell(column=11, row=iStepRow+20, value="%s" % step['mixTypeType'])
                    # Design data
                    _ = wsht.cell(column=12, row=iStepRow+20, value=s['refusalPressure'])
                    _ = wsht.cell(column=13, row=iStepRow+20, value=0)
                    _ = wsht.cell(column=14, row=iStepRow+20, value=s['minFlowRate'])
                    _ = wsht.cell(column=15, row=iStepRow+20, value=maxVolume)
                    _ = wsht.cell(column=16, row=iStepRow+20, value="%s" % groutMix['code'])
                    #Grouting time
                    _ = wsht.cell(column=17, row=iStepRow+20, value="%s" % startDateTime)
                    _ = wsht.cell(column=18, row=iStepRow+20, value="%s" % stopDateTime)
                    _ = wsht.cell(column=19, row=iStepRow+20, value="%s" % effTime)
                    _ = wsht.cell(column=20, row=iStepRow+20, value="%s" % elapsedTime)
                    #Grout Take                    
                    _ = wsht.cell(column=21, row=iStepRow+20, value="%s" % groutMix['code'])
                    _ = wsht.cell(column=22, row=iStepRow+20, value=step['groutTake'])
                    _ = wsht.cell(column=23, row=iStepRow+20, value=step['groutTake']/(stageLength))
                    _ = wsht.cell(column=24, row=iStepRow+20, value=cumV/stageLength)
                    #Flowrate
                    _ = wsht.cell(column=25, row=iStepRow+20, value=flowRateAvg)
                    _ = wsht.cell(column=26, row=iStepRow+20, value=flowRateMax)
                    _ = wsht.cell(column=27, row=iStepRow+20, value=flowRateFinal)
                    # Grouting - Gauge Pressure                   
                    _ = wsht.cell(column=28, row=iStepRow+20, value=pGaugeAvg)
                    _ = wsht.cell(column=26, row=iStepRow+20, value=pGaugeMax)
                    _ = wsht.cell(column=30, row=iStepRow+20, value=pGaugeFinal)
                    # Grouting - Gauge Pressure                   
                    _ = wsht.cell(column=31, row=iStepRow+20, value=pEffAvg)
                    _ = wsht.cell(column=32, row=iStepRow+20, value=pEffMax)
                    _ = wsht.cell(column=33, row=iStepRow+20, value=pEffFinal)
                    #comment                    
                    _ = wsht.cell(column=37, row=iStepRow+20, value="%s" % s["comment"])
                    """step_row = StepRow('boreholeid'=main_borehole['boreholeId'],'station'=main_borehole['stationId_Build'],'elevation'=main_borehole['topElevation_Build'],
                                                    'stage_sequence'=stSeq,'method'=s['procedureType'],'top_depth'=s['topLength'],
                                                    'top_elevation'=s['topLength'],'stage_length'=s['topLength'],'mixtype'=step['mixTypeType'],
                                                    'refusalP'=s['refusalPressure'],'minFlowRate'=s['minFlowRate'],'maxVolume'='ND','groutMix'=groutMix['code'],
                                                    'startStepTime'='ND','stopStepTime'=s['startDateTime']+datetime.timedelta(0,step['elapsedTime']),'elapsedTime'=step['elapsedTime'],'productionTime'='ND',
                                                    'groutTake'=step['groutTake'], 'flowRateAvg'=flowRateAvg, 'flowRateMax'=flowRateMax, 'flowRateFinal'=flowRateFinal,
                                                     'pGaugeAvg'=pGaugeAvg, 'pGaugeMax'=pGaugeMax, 'pGaugeFinal'=pGaugeFinal,'appLugeonUnit'=appLugeonUnit,'gin'=gin,'pq'=pq)"""
                    export_stages.append(step_row)
                else:
                    print "\t....without series".format(main_borehole['boreholeId'])
                    log.info("\t....without series".format(main_borehole['boreholeId']))
        for ex in export_stages:
            print ex
        # Borehole Header
        wsht['O3'] = main_borehole['boreholeId']
        wsht['G4'] = "Grouting gallery"
        wsht['G5'] = "CS3"
        wsht['G6'] = bh_section_code
        wsht['G7'] = bh_line_code
        wsht['G8'] = bh_type_codes[main_borehole['type']]+ " ("+main_borehole['type']+")"
        wsht['G9'] = main_borehole['stationId_Build']
        wsht['G10'] = main_borehole['offset_Build']
        wsht['G11'] = main_borehole['inclination_Build']
        wsht['G12'] = main_borehole['azimuth_Build']
        wsht['G13'] = main_borehole['topElevation_Build']
        # TODO calcolare Depth
        wsht['G14'] = main_borehole['holeLength_Build']
        wsht['G15'] = main_borehole['topElevation_Build'] - main_borehole['holeLength_Build']

        wb.save( "xlsx/data_record_"+main_borehole['boreholeId']+".xlsx" )


if __name__ == "__main__":
    main(sys.argv[1:])
