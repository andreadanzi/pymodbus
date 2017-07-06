# -*- coding: utf-8 -*-
#!/usr/bin/env python
import pymongo
from pymongo import MongoClient
from bson.objectid import ObjectId
import logging
import logging.handlers
import os
import sys
import getopt
from datetime import datetime, timedelta
import collections
from scipy.stats import randint
import numpy as np
from collections import namedtuple
from openpyxl import load_workbook, Workbook
# sudo apt-get install libgdal-dev
# sudo apt-get install python-gdal
import ConfigParser
StepRow = namedtuple('StepRow',['area','section','boreholeid','station','elevation',
                                'stage_sequence','method','top_length',
                                'bottom_length','stage_length','mixtype',
                                'refusalP','minFlowRate','maxVolume','groutMix',
                                'startStepTime','stopStepTime','elapsedTime','productionTime',
                                'groutTake', 'flowRateAvg', 'flowRateMax', 'flowRateFinal',
                                 'pGaugeAvg', 'pGaugeMax','qAtpGaugeMax', 'pGaugeFinal','pEffAvg', 'pEffMax','qAtpEffMax', 'pEffFinal','appLugeonUnit','gin','pq','solidContent','stepStatus','rCheckDuration'
                                 ,'cumVol','cumSol'
                                ])
"""
	3	2	1	k
Q	0.0001447	-0.024499	0.86506	15.4763
P	0.000021167	-0.00403	0.30249	-0.8024
"""
sCFGName = "{0}.cfg".format(os.path.basename(__file__).split(".")[0])

bh_type_codes = {"E":"Exploratory","P":"Primary","S":"Secondary","T":"Ternary","Q":"Quanternary","L":"Quinary"}

def pOut(t):
    retV=0.000021167*t**3-0.00403*t**2+0.30249*t-0.8024
    if retV < 0.0:
        retV = 0.0
    return retV

def qOut(t):
    retV = 0.0001447*t**3-0.024499*t**2+0.86506*t+15.4763
    return retV

# client = MongoClient('localhost', 27017)
log = logging.getLogger()
log.setLevel(logging.DEBUG)
file_handler = logging.handlers.RotatingFileHandler("{0}.log".format(os.path.basename(__file__).split(".")[0]), maxBytes=5000000,backupCount=5)
file_handler.setLevel(logging.DEBUG)
formatter = logging.Formatter('%(asctime)s %(name)-12s %(levelname)-8s %(message)s')
file_handler.setFormatter(formatter)
log.addHandler(file_handler)

delta_rand = randint(-2, 2)
# dati di partenza
# da caricare con
# >mongoimport --db import --collection timeseries --host srv3.sws-digital.com --file import_series.json
# generati con
# >mongoexport --db tgrout-development --collection timeseries --host srv3.sws-digital.com --out import_series.json --query "{'stage':ObjectId('57c6cbb9ea71da9c2513d5b2'),'step':ObjectId('57c6cbb9ea71da9c2513d5b3')}"
# 57cd5367deeb0e074430c8f7
# 57cd5367deeb0e074430c8f8
import_refusalPressure = 10.
base_stageId = "57c6cbb9ea71da9c2513d5b2"
base_stepId = "57c6cbb9ea71da9c2513d5b3"



# python export_xlsx_steps.py -m localhost -p 27017 -d tgrout-dev -b 57ab48f1f194b0873319a12a
def main(argv):
    syntax = "python " + os.path.basename(__file__) + " -m <mongo host> -p <mongo port> -d <main database> -b <borehole_id> -s <sectioncode> -c"
    mongo_host = "localhost"
    mongo_port = 27017
    mongo_database = "tgrout-development"
    sCurrentWorkingdir = os.getcwd()
    sDate = datetime.utcnow().strftime("%Y%m%d%H%M%S")    
    splitsections = False
    sectionCode = 0
    borehole_id = None
    bUseConfig = False
    project_code = "P2016_015-MOSUL"
    root_folder = "/home/andrea/"
    try:
        opts = getopt.getopt(argv, "hm:p:d:b:s:c", ["mongohost=","port=","database=","borehole_id=","sectioncode=","config"])[0]
    except getopt.GetoptError:
        print syntax
        sys.exit(1)
    if len(opts) < 1:
        print syntax
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            print syntax
            sys.exit()
        elif opt in ("-p", "--port"):
            mongo_port = int(arg)
        elif opt in ("-s", "--sectioncode"):
            sectionCode = int(arg)
        elif opt in ("-c", "--config"):
            bUseConfig = True
        elif opt in ("-m", "--mongohost"):
            mongo_host = arg
        elif opt in ("-d", "--database"):
            mongo_database = arg
        elif opt in ("-b", "--borehole_id"):
            borehole_id = arg
    
    if bUseConfig:
        if os.path.exists(sCFGName):    
            syncConfig = ConfigParser.RawConfigParser()
            cfgItems = syncConfig.read(sCFGName)
            mongo_host = syncConfig.get('MongoDB', 'host')    
            mongo_database = syncConfig.get('MongoDB', 'db')    
            mongo_port = int(syncConfig.get('MongoDB', 'port')) 
            project_code = syncConfig.get('Project', 'code')     
            root_folder = syncConfig.get('System', 'root_folder')
        else:
            errMsg = "Error: config file {0} does not exists!".format(sCFGName)
            log.error(errMsg)
            print errMsg
            sys.exit()


    project_folder = os.path.join(root_folder, "projects", project_code)  
    gis_folder = os.path.join(project_folder, "gis")    
    
    if os.path.exists(gis_folder):
        log.debug("GIS folder {0} exists".format(gis_folder))
    else:
        log.info("{0} does not exists".format(gis_folder))
        os.makedirs(gis_folder)
        log.info("{0} created".format(gis_folder))
    
    
    
    mClient = MongoClient(mongo_host, mongo_port)
    db = mClient[mongo_database]
    
        
    headLossFactorCache={}
    """
    all_constructionsites = list(db.constructionsites.find())
    for csite_item in all_constructionsites:
    """
    if sectionCode > 0:
        secFilter = {"code":sectionCode}
    else:
        secFilter = {}
    all_sections = list(db.sections.find().sort('code', pymongo.ASCENDING))
    for n_sect, section_item in enumerate(all_sections):
        csite_item = db.constructionsites.find_one({"_id":section_item["constructionSite"]})
        if csite_item:
            export_stages = []
            scCode = section_item["code"]
            print scCode
            fromBh = None
            toBh = None
            sectionVolume = 0.
            sectionSolidContent = 0.
            all_lines = db.lines.find({"section":section_item["_id"]})
          
            #
            for line_item in all_lines:
                all_boreholes = list(db.boreholes.find({"constructionSite":csite_item["_id"],"section":section_item["_id"],"line":line_item["_id"]}).sort('position', pymongo.ASCENDING))
                bh_line = line_item
                for enum, main_borehole in enumerate(all_boreholes):
                    stages = db.stages.find({"borehole":main_borehole["_id"], "refusalPressure":{"$exists":True}, "rCheckDuration":{"$exists":True}}).sort('startDateTime', pymongo.ASCENDING)

                    if fromBh == None:
                        fromBh = main_borehole                                       
                    toBh = main_borehole
                    
                    if "location" not in main_borehole:
                        print "Missing location for {}".main_borehole["borehole_id"]
                    log.info("constructionSite {5} => Found borehole with Id {0}! topElevation_Build = {1}, topElevation_Build={2}, holeLength_Design={3} and holeLength_Design={4}".format(main_borehole["_id"],main_borehole["topElevation_Build"],main_borehole["topElevation_Build"],main_borehole["holeLength_Design"],main_borehole["holeLength_Design"],main_borehole["constructionSite"]))        
                    # boreholeId boreholeId
                    # section
                    # line
                    # type
                    # position
                    # stationId_Build
                    # topElevation_Build
                    # offset_Build
                    # inclination_Build
                    # azimuth_Build
                    # print main_borehole["boreholeId"]
                    bh_totVolume = 0.
                    bh_totSolidContent = 0.
                    stgs = list(stages)
                    for stSeq, s in enumerate( stgs ):
                        log.info("ID {0} stageStatus {1}".format(s["_id"],s["stageStatus"]))
                        cumV = 0.
                        solidContent = 0.
                        finalP = 0.
                        finalQ = 0.
                        MaxP = 0.
                        MaxQ = 0.  
                        for ist, step in enumerate(s["steps"]):
                            if "mixTypeType" not in step:
                                print step["_id"]
                                continue
                            rCheckDuration = s["rCheckDuration"]
                            log.info("\tID {0} stepStatus {1}".format(step["_id"],step["stepStatus"]))
                            timeseries = db.timeseries.find({"stage":ObjectId(s["_id"]),"step":ObjectId(step["_id"])}).sort('timestampMinute', pymongo.ASCENDING)
                            
                            groutmix = None
                            sHdlf = str(step["headLossFactor"])
                            if sHdlf in headLossFactorCache:
                                groutmix = headLossFactorCache[ sHdlf ]
                            else:
                                headLossFactor = db.headlossfactors.find_one({"_id":ObjectId(step["headLossFactor"])})
                                groutmix = db.mixtypes.find_one({"_id":ObjectId(headLossFactor["mixType"])})
                                headLossFactorCache[ sHdlf ] = groutmix
                                cementWater = groutmix["cementWater"]
                                bentoniteWater = groutmix["bentoniteWater"]
                                sandWater = groutmix["sandWater"]
                                waterGs = groutmix["waterGs"]
                                cementGs = groutmix["cementGs"]
                                bentoniteGs = groutmix["bentoniteGs"]
                                sandGs = groutmix["sandGs"]
                                groutmix["solidRate"] = (cementWater+bentoniteWater+sandWater)/(1/waterGs + cementWater/cementGs + bentoniteWater/bentoniteGs + sandWater/sandGs  )
                                
                                # print "Solid Rate {0}={1}".format(groutmix["code"],groutmix["solidRate"])
                                headLossFactorCache[ sHdlf ] = groutmix
                            
                            lastPe = collections.deque(maxlen=rCheckDuration*60)
                            lastPg = collections.deque(maxlen=rCheckDuration*60)
                            lastQ = collections.deque(maxlen=rCheckDuration*60)
                            avgPe = 0.
                            avgPg = 0.
                            avgQ = 0.
                            avg3Pe = [0.,0.,0.]
                            avg3Pg = [0.,0.,0.]
                            avg3Q = [0.,0.,0.]
                            maxPe = 0.
                            maxPg = 0.
                            maxQ = 0.
                            qAtMaxPe = 0.
                            qAtMaxPg = 0.
                            iMaxAvg = 3
                            
                            iCount = 0
                            if timeseries.count():
                                startDateTime = timeseries[0]['timestampMinute']
                                for tv in list(timeseries):
                                    stopDateTime = tv['timestampMinute']
                                    lastSeconds = 0
                                    tv_values = [tval for tval in tv["values"].items() if tval[0] not in ('_id','v')  ]
                                    tvo = collections.OrderedDict(sorted(tv_values, key=lambda t: int(t[0])))
                                    for vk in tvo:
                                        if "avgPressureEffective" in tvo[vk]:
                                            lastSeconds += 1
                                            iCount = iCount + 1
                                            """
                                            "flowRateGaugeManifold" : 17.9745810337963,
                                            "pressureGaugeManifold" : 0.625024111325,
                                            "pressureEffective" : 0.520853426104166,
                                            "elapsedTime" : 57000,
                                            "_id" : ObjectId("57d2c23177c87c26d6a06971"),
                                            "injectedGroutTakeVolume" : 15.9477824093287
                                            """
                                            lastQ.append(tvo[vk]["flowRateGaugeManifold"])
                                            lastPe.append(tvo[vk]["pressureEffective"])
                                            lastPg.append(tvo[vk]["pressureGaugeManifold"])
                                            avgPe = avgPe + (tvo[vk]["pressureEffective"] - avgPe)/(iCount+1)
                                            avgPg = avgPg + (tvo[vk]["pressureGaugeManifold"] - avgPg)/(iCount+1)
                                            avgQ = avgQ + (tvo[vk]["flowRateGaugeManifold"] - avgQ)/(iCount+1)
                                            avg3Pe[(iCount-1)%iMaxAvg] = tvo[vk]["pressureEffective"] 
                                            avg3Pg[(iCount-1)%iMaxAvg] = tvo[vk]["pressureGaugeManifold"] 
                                            avg3Q[(iCount-1)%iMaxAvg] = tvo[vk]["flowRateGaugeManifold"] 
                                            
                                            if (iCount-1)%iMaxAvg == iMaxAvg-1:
                                                if np.mean(avg3Pe) > maxPe:
                                                    maxPe = np.mean(avg3Pe)
                                                    qAtMaxPe = np.mean(avg3Q)
                                                if np.mean(avg3Pg) > maxPg:
                                                    maxPg = np.mean(avg3Pg)
                                                    qAtMaxPg = np.mean(avg3Q)
                                                if np.mean(avg3Q) > maxQ:
                                                    maxQ = np.mean(avg3Q)
                                        else:
                                            pass
                                
                                stopDateTime = stopDateTime + timedelta(seconds=lastSeconds)                    
                                elapsedTime = stopDateTime - startDateTime
                                stageLength = s['bottomLength'] - s['topLength']
                                flowRateAvg = avgQ/stageLength
                                flowRateMax = maxQ/stageLength
                                lastQ4Mean = [dt for cnf, dt in enumerate(lastQ) if (cnf+1) % 10 ==0 ]
                                lastPe4Mean = [dt for cnf, dt in enumerate(lastPe) if (cnf+1) % 10 ==0 ]
                                lastPg4Mean = [dt for cnf, dt in enumerate(lastPg) if (cnf+1) % 10 ==0 ]
                                flowRateFinal = float('nan')
                                if len(lastQ4Mean) > 0:
                                    flowRateFinal = np.mean(lastQ4Mean)/stageLength
                                pGaugeAvg = avgPg
                                pGaugeMax = maxPg
                                qAtpGaugeMax = qAtMaxPg/stageLength
                                pGaugeFinal = float('nan')
                                if len(lastPg4Mean) > 0:
                                    pGaugeFinal = np.mean(lastPg4Mean)
                                pEffAvg = avgPe
                                pEffMax = maxPe
                                qAtpEffMax = qAtMaxPe/stageLength
                                if pEffMax > MaxP:
                                    MaxP = pEffMax
                                    MaxQ = qAtpEffMax
                                pEffFinal = float('nan')
                                if len(lastPe4Mean) > 0:
                                    pEffFinal = np.mean(lastPe4Mean)
                                    appLugeonUnit = 10.*flowRateFinal/pEffFinal
                                gin = step['groutTake']*pEffFinal/stageLength
                                if flowRateFinal > 0:
                                    pq = pEffFinal/flowRateFinal
                                effTime = timedelta(seconds=iCount)
                                maxVolume = 99
                                cumV += step['groutTake']
                                solidContent += groutmix["solidRate"]*step['groutTake']
                                #A21  AJ21                  
                                step_row = StepRow(csite_item["code"], scCode,main_borehole['boreholeId'], #A21+i
                                                    main_borehole['stationId_Build'],
                                                    main_borehole['topElevation_Build'],
                                                    stSeq + 1,
                                                    s['procedureType'],
                                                    s['topLength'],
                                                    s['bottomLength'],
                                                    s['bottomLength'] - s['topLength'],
                                                    step['mixTypeType'],
                                                    s['refusalPressure'],
                                                    s['minFlowRate'],
                                                    maxVolume,
                                                    groutmix['code'],
                                                    str(startDateTime),
                                                    str(stopDateTime),
                                                    str(elapsedTime),
                                                    str(effTime),
                                                    step['groutTake'],
                                                    flowRateAvg,
                                                    flowRateMax,
                                                    flowRateFinal,
                                                    pGaugeAvg,
                                                    pGaugeMax,
                                                    qAtpGaugeMax,
                                                    pGaugeFinal,
                                                    pEffAvg,
                                                    pEffMax,
                                                    qAtpEffMax,
                                                    pEffFinal,
                                                    appLugeonUnit,
                                                    gin,
                                                    pq,
                                                    groutmix["solidRate"]*step['groutTake'],
                                                    step["stepStatus"],
                                                    s['rCheckDuration'],
                                                    cumV,
                                                    solidContent)
                                export_stages.append(step_row)
                                
                                finalP = pEffFinal
                                finalQ = flowRateFinal
                            else:
                                print "BH {0} Stage ID {1} step {2} stepStatus {3} without series".format(main_borehole['boreholeId'], s["_id"],step["_id"], step["stepStatus"])
                                log.info("\t{0}....without series".format(main_borehole['boreholeId']))   
                            
                        bh_totVolume += cumV
                        sectionVolume += cumV
                        bh_totSolidContent += solidContent
                        sectionSolidContent += solidContent
            if len(export_stages) > 0:
                wb = Workbook()
                wsht = wb.create_sheet(title="Data")
                iStepRow = 1
                """
                ['area','section','boreholeid','station','elevation',
                                            'stage_sequence','method','top_length',
                                            'bottom_length','stage_length','mixtype',
                                            'refusalP','minFlowRate','maxVolume','groutMix',
                                            'startStepTime','stopStepTime','elapsedTime','productionTime',
                                            'groutTake', 'flowRateAvg', 'flowRateMax', 'flowRateFinal',
                                             'pGaugeAvg', 'pGaugeMax','qAtpGaugeMax', 'pGaugeFinal','pEffAvg', 'pEffMax','qAtpEffMax', 'pEffFinal','appLugeonUnit','gin','pq'
                                            ]
                """
                 # Borehole data
                _ = wsht.cell(column=1, row=iStepRow, value="BOREHOLE ID")
                _ = wsht.cell(column=2, row=iStepRow, value="STATION")
                _ = wsht.cell(column=3, row=iStepRow, value= "AREA")
                _ = wsht.cell(column=4, row=iStepRow, value= "ELEVATION")
                # Stage data
                _ = wsht.cell(column=5, row=iStepRow, value="SEQUENCE")
                _ = wsht.cell(column=6, row=iStepRow, value="METHOD")
            
                _ = wsht.cell(column=7, row=iStepRow, value="STEP_STATUS") #"Stage (top-down)"
                _ = wsht.cell(column=8, row=iStepRow, value="TOP LENGTH") # Top Depth
                _ = wsht.cell(column=9, row=iStepRow, value="BOTTOM LENGTH")
                _ = wsht.cell(column=10, row=iStepRow, value="STAGE LENGTH")
                _ = wsht.cell(column=11, row=iStepRow, value="STEP TYPE")
                # Design data
                _ = wsht.cell(column=12, row=iStepRow, value="R (design)")
                _ = wsht.cell(column=13, row=iStepRow, value="RTW")
                _ = wsht.cell(column=14, row=iStepRow, value="MIN FlowRate (design)")
                _ = wsht.cell(column=15, row=iStepRow, value="MAX Volume (design)")
                _ = wsht.cell(column=16, row=iStepRow, value="Grout MIX")
                #Grouting time
                _ = wsht.cell(column=17, row=iStepRow, value="START")
                _ = wsht.cell(column=18, row=iStepRow, value="STOP")
                _ = wsht.cell(column=19, row=iStepRow, value="Grouting Time")
                _ = wsht.cell(column=20, row=iStepRow, value="Elapsed Time")
                #Grout Take                    
                _ = wsht.cell(column=21, row=iStepRow, value="Grout Take")
                _ = wsht.cell(column=22, row=iStepRow, value="Grout Take/m")
                _ = wsht.cell(column=23, row=iStepRow, value="Solid Content")
                _ = wsht.cell(column=24, row=iStepRow, value="Solid Content/m")
                #Flowrate
                _ = wsht.cell(column=25, row=iStepRow, value="Flow Rate Avg")
                _ = wsht.cell(column=26, row=iStepRow, value="Flow Rate Max")
                _ = wsht.cell(column=27, row=iStepRow, value="Flow Rate Final")
                # Grouting - Gauge Pressure                   
                _ = wsht.cell(column=28, row=iStepRow, value="P Gauge Avg")
                _ = wsht.cell(column=29, row=iStepRow, value="P Gauge Max")
                _ = wsht.cell(column=30, row=iStepRow, value="P Gauge Final")
                # Grouting - Gauge Pressure                   
                _ = wsht.cell(column=31, row=iStepRow, value="P Eff Avg")
                _ = wsht.cell(column=32, row=iStepRow, value="P Eff Max")
                _ = wsht.cell(column=33, row=iStepRow, value="P Eff Final")
                # Grouting - Gauge Pressure                   
                _ = wsht.cell(column=34, row=iStepRow, value="Tot. Grout Take")
                _ = wsht.cell(column=35, row=iStepRow, value="Tot. Solid Content")
                iStepRow += 1
                #comment       
                for item in export_stages:
                     # Borehole data
                    _ = wsht.cell(column=1, row=iStepRow, value="%s" % item.boreholeid)
                    _ = wsht.cell(column=2, row=iStepRow, value= float(item.station))
                    _ = wsht.cell(column=3, row=iStepRow, value= "%s" % item.area)
                    _ = wsht.cell(column=4, row=iStepRow, value= item.elevation)
                    # Stage data
                    _ = wsht.cell(column=5, row=iStepRow, value=item.stage_sequence)
                    _ = wsht.cell(column=6, row=iStepRow, value="%s" % item.method)
            
                    _ = wsht.cell(column=7, row=iStepRow, value="%s" % item.stepStatus) #"Stage (top-down)"stageStatus
                    _ = wsht.cell(column=8, row=iStepRow, value=item.top_length) # Top Depth
                    _ = wsht.cell(column=9, row=iStepRow, value=item.bottom_length)
                    _ = wsht.cell(column=10, row=iStepRow, value=item.stage_length)
                    _ = wsht.cell(column=11, row=iStepRow, value="%s" % item.mixtype)
                    # Design data
                    _ = wsht.cell(column=12, row=iStepRow, value=item.refusalP)
                    _ = wsht.cell(column=13, row=iStepRow, value=item.rCheckDuration)
                    _ = wsht.cell(column=14, row=iStepRow, value=item.minFlowRate)
                    _ = wsht.cell(column=15, row=iStepRow, value=item.maxVolume)
                    _ = wsht.cell(column=16, row=iStepRow, value="%s" % item.groutMix)
                    #Grouting time
                    _ = wsht.cell(column=17, row=iStepRow, value="%s" % item.startStepTime)
                    _ = wsht.cell(column=18, row=iStepRow, value="%s" % item.stopStepTime)
                    _ = wsht.cell(column=19, row=iStepRow, value="%s" % item.productionTime)
                    _ = wsht.cell(column=20, row=iStepRow, value="%s" % item.elapsedTime)
                    #Grout Take                    
                    _ = wsht.cell(column=21, row=iStepRow, value=item.groutTake)
                    _ = wsht.cell(column=22, row=iStepRow, value=item.groutTake/item.stage_length)
                    _ = wsht.cell(column=23, row=iStepRow, value=item.solidContent)
                    _ = wsht.cell(column=24, row=iStepRow, value=item.solidContent/item.stage_length)
                    #Flowrate
                    _ = wsht.cell(column=25, row=iStepRow, value=item.flowRateAvg)
                    _ = wsht.cell(column=26, row=iStepRow, value=item.flowRateMax)
                    _ = wsht.cell(column=27, row=iStepRow, value=item.flowRateFinal)
                    # Grouting - Gauge Pressure                   
                    _ = wsht.cell(column=28, row=iStepRow, value=item.pGaugeAvg)
                    _ = wsht.cell(column=29, row=iStepRow, value=item.pGaugeMax)
                    _ = wsht.cell(column=30, row=iStepRow, value=item.pGaugeFinal)
                    # Grouting - Gauge Pressure                   
                    _ = wsht.cell(column=31, row=iStepRow, value=item.pEffAvg)
                    _ = wsht.cell(column=32, row=iStepRow, value=item.pEffMax)
                    _ = wsht.cell(column=33, row=iStepRow, value=item.pEffFinal)
                    _ = wsht.cell(column=34, row=iStepRow, value=item.cumVol)
                    _ = wsht.cell(column=35, row=iStepRow, value=item.cumSol)
                    #comment       
            
                    iStepRow += 1
                wb.save("export_section_{0:03d}.xlsx".format(scCode))  
       
 

        


if __name__ == "__main__":
    main(sys.argv[1:])
